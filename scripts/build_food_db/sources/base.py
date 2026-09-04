import csv
from typing import Dict, Iterator, List, Optional

from scripts.build_food_db.model import NormalisedRow


class Source:
    """Base class for a national / regional food-composition extractor.

    Subclasses set ``id`` and implement ``extract(raw_dir)``, returning a
    list of :class:`NormalisedRow` already run through ``normalise_row``.
    """

    id = ""

    def extract(self, raw_dir: str) -> List[NormalisedRow]:
        raise NotImplementedError


def to_mg(value_g: Optional[float]) -> Optional[float]:
    """Grams -> milligrams. ``None`` passes through."""
    return None if value_g is None else round(value_g * 1000.0, 4)


def to_mcg(value_g: Optional[float]) -> Optional[float]:
    """Grams -> micrograms. ``None`` passes through."""
    return None if value_g is None else round(value_g * 1_000_000.0, 4)


def kj_to_kcal(value_kj: Optional[float]) -> Optional[float]:
    """Kilojoules -> kilocalories (1 kcal = 4.184 kJ). ``None`` passes through."""
    return None if value_kj is None else round(value_kj / 4.184, 4)


def read_csv_rows(path: str, delimiter: str = ",",
                  encoding: str = "utf-8") -> Iterator[Dict[str, str]]:
    with open(path, newline="", encoding=encoding) as fh:
        yield from csv.DictReader(fh, delimiter=delimiter)


def read_xlsx_rows(path: str, sheet: Optional[str] = None) -> Iterator[Dict[str, str]]:
    """Yield dict rows from an .xlsx sheet, keyed by the first row's headers.

    Uses ``openpyxl`` in ``read_only=True`` mode so multi-thousand-row
    national tables stream instead of loading whole. ``sheet`` may be a
    sheet name (str); ``None`` uses the active sheet.
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet] if sheet is not None else wb.active
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            return
        headers = ["" if h is None else str(h).strip() for h in header]
        for raw in rows:
            record = {}
            for idx, key in enumerate(headers):
                if key == "":
                    continue
                value = raw[idx] if idx < len(raw) else None
                record[key] = "" if value is None else str(value)
            yield record
    finally:
        wb.close()


_NULL_TOKENS = {
    "", "-", "n", "tr", "trace", "traces", "[n]", "nd", "n/a", "na", "*",
}


def parse_float(raw: Optional[str]) -> Optional[float]:
    if raw is None:
        return None
    text = raw.strip().replace("\xa0", "").replace(" ", "")
    # A comma is a decimal separator (French / Danish tables: "12,34") only
    # when there is exactly one and at most two digits follow it. Otherwise it
    # is a thousands separator ("1,200" -> 1200) and is stripped.
    if text.count(",") == 1:
        intpart, frac = text.split(",")
        if 0 < len(frac) <= 2 and frac.isdigit():
            text = intpart + "." + frac
        else:
            text = text.replace(",", "")
    else:
        text = text.replace(",", "")
    # national tables mark "below detection limit" as "< 0.1" etc.
    if text.startswith("<"):
        return None
    if text.lower() in _NULL_TOKENS:
        return None
    try:
        return float(text)
    except ValueError:
        return None
