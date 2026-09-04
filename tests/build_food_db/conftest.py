import csv
import os
import shutil

FIX = os.path.join(os.path.dirname(__file__), "fixtures")


def csv_to_xlsx(csv_path, xlsx_path, sheet=None, extra_rows=()):
    """Write a committed CSV fixture out as a one-sheet .xlsx workbook.

    The xlsx sources are fed real workbooks, but a CSV fixture stays readable
    and diffable in git — so the workbook is built at test setup time rather
    than committing a binary blob. ``extra_rows`` are appended after the CSV.
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    if sheet is not None:
        ws.title = sheet
    with open(csv_path, newline="", encoding="utf-8") as fh:
        for row in csv.reader(fh):
            ws.append(row)
    for row in extra_rows:
        ws.append(list(row))
    wb.save(str(xlsx_path))
    return str(xlsx_path)


def cofid_raw_dir(tmp_path, csv_path=None, extra_rows=()):
    """Build the CoFID workbook the extractor expects: one workbook with
    the three real sheets (Proximates / Inorganics / Vitamins) it joins on
    the first column, split out of the single committed CSV slice.
    ``extra_rows`` are positional value lists matching that CSV's own
    header order, appended after its rows."""
    from openpyxl import Workbook
    from scripts.build_food_db.sources import cofid

    d = tmp_path / "cofid"
    if not d.exists():
        d.mkdir(parents=True)

    with open(csv_path or os.path.join(FIX, "cofid_slice.csv"), newline="", encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        fieldnames = reader.fieldnames
        rows = list(reader)
    for extra in extra_rows:
        rows.append(dict(zip(fieldnames, extra)))

    prox_cols = ["Food Code", "Food Name", "Description", "Group",
                 "Water (g)", "Protein (g)", "Fat (g)", "Carbohydrate (g)",
                 "Energy (kcal) (kcal)", "Energy (kJ) (kJ)",
                 "Total sugars (g)", "AOAC fibre (g)"]
    inorg_cols = ["Food Code", "Food Name", "Sodium (mg)", "Calcium (mg)", "Iron (mg)"]
    vit_cols = ["Food Code", "Food Name", "Vitamin C (mg)", "Vitamin D (µg)", "Vitamin B12 (µg)"]

    wb = Workbook()
    wb.active.title = cofid.SHEET_PROXIMATES
    for sheet_name, cols in (
        (cofid.SHEET_PROXIMATES, prox_cols),
        (cofid.SHEET_INORGANICS, inorg_cols),
        (cofid.SHEET_VITAMINS, vit_cols),
    ):
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
        ws.append(cols)
        for row in rows:
            ws.append([row.get(c, "") for c in cols])
    wb.save(str(d / cofid.FILE))
    return str(d)


def usda_raw_dir(tmp_path):
    """Copy the committed USDA *_slice.csv fixtures into tmp_path/usda/
    under the real FDC release filenames, and return that dir as a str."""
    d = tmp_path / "usda"
    d.mkdir()
    shutil.copy(os.path.join(FIX, "usda_food_slice.csv"), d / "food.csv")
    shutil.copy(os.path.join(FIX, "usda_food_nutrient_slice.csv"),
                d / "food_nutrient.csv")
    shutil.copy(os.path.join(FIX, "usda_nutrient_slice.csv"), d / "nutrient.csv")
    return str(d)


def cnf_raw_dir(tmp_path):
    """Copy the committed CNF *_slice.csv fixtures into tmp_path/cnf/ under
    the real 2026 full-file release names, and return that dir."""
    d = tmp_path / "cnf"
    d.mkdir()
    shutil.copy(os.path.join(FIX, "cnf_food_name_slice.csv"), d / "food_name.csv")
    shutil.copy(os.path.join(FIX, "cnf_nutrient_name_slice.csv"),
                d / "nutrient_name.csv")
    shutil.copy(os.path.join(FIX, "cnf_nutrient_amount_slice.csv"),
                d / "nutrient_amount.csv")
    return str(d)


def afcd_raw_dir(tmp_path):
    """Copy the committed AFCD *_slice.xlsx fixtures into tmp_path/afcd/ under
    the real Release 3 workbook names, and return that dir."""
    d = tmp_path / "afcd"
    d.mkdir()
    shutil.copy(os.path.join(FIX, "afcd_food_details_slice.xlsx"),
                d / "AFCD Release 3 - Food Details.xlsx")
    shutil.copy(os.path.join(FIX, "afcd_nutrients_slice.xlsx"),
                d / "AFCD Release 3 - Nutrient profiles.xlsx")
    return str(d)


def single_file_raw_dir(tmp_path, source_id, fixture_name, real_name):
    """Copy one committed fixture slice into tmp_path/<source_id>/<real_name>
    and return that dir as a str. Used by the single-file extractors."""
    d = tmp_path / source_id
    d.mkdir()
    shutil.copy(os.path.join(FIX, fixture_name), d / real_name)
    return str(d)
